"""Excel / CSV 报价历史导入(excel_* 源):一次性快照进落地层,同一下游。

工作流(Quotation excel binding notes 的落地):
1. excel-suggest:读表头,启发式建议 列→属性 映射,输出带置信度注释的 YAML
   ——"AI 建议"的开源版为启发式匹配;将来接 LLM 建议,产物仍是这份映射文件;
2. 人工确认 / 修订映射文件(一次确认,长期记住);
3. excel-import:按映射落地为 raw_{source}__{锚表}(列名 = 对象属性,值原样),
   缺业务键的行跳过并逐行报告;
4. 标准 apply 物化对象层 —— 校验 / 隔离区 / 熔断对 Excel 数据原样生效。

值翻译(如工厂写"中标"而对象模型是"成交")不在导入层做,写进 binding
field_map 的 (map 源值→对象值) —— 元模型保持唯一事实来源。
CSV 用标准库;.xlsx 需 excel 依赖组(openpyxl,惰性导入)。
"""

from __future__ import annotations

import csv
import difflib
import uuid
from dataclasses import dataclass, field
from pathlib import Path

import yaml

from ..mapping import parse_field_expr
from ..metamodel.schema import ObjectTemplate
from .adapters.base import TableInfo
from .landing import LandingStore

_PORTABLE = {"int": "int", "decimal": "real", "money": "real"}  # 其余落 text
_MIN_SCORE = 0.5


# ---- 读取(CSV / XLSX)----

def read_tabular(path: str | Path, sheet: str | None = None,
                 header_row: int = 1) -> tuple[list[str], list[dict]]:
    """返回 (表头, 行 dict 列表);全空行剔除,单元格字符串两端去空白。"""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as f:
            raw = list(csv.reader(f))
    elif suffix in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook  # 惰性导入:excel 依赖组
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        raw = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        raise ValueError(f"不支持的文件类型 '{suffix}'(支持 .csv / .xlsx / .xlsm)")

    if len(raw) < header_row:
        raise ValueError(f"文件行数不足,header_row={header_row}")
    headers = [str(c).strip() if c is not None else "" for c in raw[header_row - 1]]

    def norm(v):
        if isinstance(v, str):
            v = v.strip()
        return None if v in ("", None) else v

    rows = []
    for r in raw[header_row:]:
        cells = [norm(c) for c in r]
        if any(c is not None for c in cells):
            rows.append(dict(zip(headers, cells)))
    return headers, rows


# ---- 列映射建议(启发式)----

def _desc_base(desc: str) -> str:
    for sep in ("(", "("):
        desc = desc.split(sep)[0]
    return desc.strip()


def _score(header: str, prop) -> float:
    h = header.strip()
    if not h:
        return 0.0
    if h.lower() == prop.name.lower():
        return 1.0
    desc = (prop.desc or "").strip()
    if h == _desc_base(desc):
        return 0.95
    if desc and h in desc:
        return 0.85
    ratio = max(difflib.SequenceMatcher(None, h, _desc_base(desc)).ratio(),
                difflib.SequenceMatcher(None, h.lower(), prop.name).ratio())
    return ratio * 0.7


def suggest_mapping(headers: list[str], tpl: ObjectTemplate) -> dict:
    """贪心分配:{columns: {表头: 属性}, scores: {表头: 分}, unmapped: [表头]}。"""
    pairs = sorted(
        ((header, p.name, _score(header, p))
         for header in headers if header for p in tpl.properties),
        key=lambda x: -x[2])
    columns: dict[str, str] = {}
    scores: dict[str, float] = {}
    taken: set[str] = set()
    for header, prop, score in pairs:
        if score < _MIN_SCORE or header in columns or prop in taken:
            continue
        columns[header] = prop
        scores[header] = round(score, 2)
        taken.add(prop)
    unmapped = [h for h in headers if h and h not in columns]
    return {"columns": columns, "scores": scores, "unmapped": unmapped}


def render_mapping_yaml(tpl: ObjectTemplate, suggestion: dict,
                        sheet: str | None = None, header_row: int = 1) -> str:
    lines = [
        "# excel-suggest 生成的列映射建议:人工确认 / 修订后供 excel-import 使用",
        f"# 对象属性全集:{[p.name for p in tpl.properties]}",
    ]
    if suggestion["unmapped"]:
        lines.append(f"# ⚠ 未映射表头(确认后手工补进 columns):{suggestion['unmapped']}")
    lines += [
        f"object: {tpl.object}",
        f"sheet: {sheet if sheet is not None else 'null'}",
        f"header_row: {header_row}",
        "columns:",
    ]
    for header, prop in suggestion["columns"].items():
        lines.append(f"  {header}: {prop}    # 置信度 {suggestion['scores'][header]}")
    return "\n".join(lines) + "\n"


def load_mapping(path: str | Path) -> dict:
    data = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    for key in ("object", "columns"):
        if not data.get(key):
            raise ValueError(f"映射文件缺少 '{key}'(由 excel-suggest 生成后人工确认)")
    return data


# ---- 导入 ----

@dataclass
class ExcelImportReport:
    source: str
    table: str
    file: str
    total: int
    imported: int
    batch_id: str
    skipped: list[tuple[int, str]] = field(default_factory=list)  # (Excel 行号, 原因)


def import_tabular(landing: LandingStore, tpl: ObjectTemplate, source: str,
                   path: str | Path, columns: dict[str, str],
                   sheet: str | None = None, header_row: int = 1) -> ExcelImportReport:
    props = {p.name: p for p in tpl.properties}
    bad = [p for p in columns.values() if p not in props]
    if bad:
        raise ValueError(f"映射到了未知属性 {bad},可用:{sorted(props)}")
    missing_keys = [k for k in tpl.keys if k not in set(columns.values())]
    if missing_keys:
        raise ValueError(f"业务键 {missing_keys} 未被任何表头映射,无法幂等导入")

    binding = next((b for b in tpl.bindings if b.source == source and b.enabled), None)
    if binding is None or not binding.tables:
        raise ValueError(f"{tpl.object} 没有 source={source} 的可用 binding(或未声明锚表)")
    anchor = binding.tables[0]

    # raw 表结构由 binding 契约决定(而非本次文件恰好有哪些列):
    # 文件缺列 → 落 NULL;映射到 binding 之外的属性 → 报错引导先补 binding
    exprs = {p: parse_field_expr(v) for p, v in binding.field_map.items()}
    anchor_col = {p: e.column for p, e in exprs.items() if e.table == anchor}
    outside = [p for p in columns.values() if p not in anchor_col]
    if outside:
        raise ValueError(
            f"属性 {outside} 不在 binding({source})的锚表 field_map 中;"
            "请先在模板 binding 里补充映射,再导入")
    info = TableInfo(
        name=anchor,
        columns=[(anchor_col[p], _PORTABLE.get(props[p].type, "text"))
                 for p in anchor_col],
        pk=[anchor_col[k] for k in tpl.keys])
    landing.ensure_raw_table(source, info)

    headers, rows = read_tabular(path, sheet, header_row)
    unknown = [h for h in columns if h not in headers]
    if unknown:
        raise ValueError(f"映射文件里的表头 {unknown} 不在文件中,实际表头:{headers}")

    batch_id = uuid.uuid4().hex[:12]
    report = ExcelImportReport(source=source, table=anchor, file=str(path),
                               total=len(rows), imported=0, batch_id=batch_id)
    good: list[dict] = []
    for i, row in enumerate(rows):
        record = {anchor_col[prop]: row.get(header) for header, prop in columns.items()}
        row_no = header_row + 1 + i
        empty_keys = [k for k in tpl.keys if record.get(anchor_col[k]) is None]
        if empty_keys:
            report.skipped.append((row_no, f"业务键缺失:{empty_keys}"))
            continue
        good.append(record)
    report.imported = landing.upsert_rows(source, info, good, batch_id)
    return report
